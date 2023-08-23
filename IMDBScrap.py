import openpyxl
import requests
from bs4 import BeautifulSoup

url="https://www.imdb.com/chart/top/" #Url we will connect
response=requests.get(url)
### Connecting url with get method

Soup=BeautifulSoup(response.content,"html5lib")
### Creating soup object for getting html elements
List=Soup.find("tbody",{"class":"lister-list"}).find_all("tr")

wb=openpyxl.Workbook()
###Created workbook object
sheet=wb.active
###Created WorkSheet

row=1
###Defined row variable for our loop
for Movie in List:
     Name=Movie.find("td",{"class":"titleColumn"}).a.text
     Year=Movie.find("td",{"class":"titleColumn"}).span.text
     Rating=Movie.find("td",{"class":"ratingColumn imdbRating"}).text.strip()
     ###Reached the html element of movie data
     FirstC=sheet.cell(column=1,row=Row)
     SecondC=sheet.cell(column=2,row=Row)
     ThirdC=sheet.cell(column=3,row=Row)
     ###Defined columns
     FirstC.value=Name
     SecondC.value=Year
     ThirdC.value=Rating
     ###Assigned values to columns
     Row+=1

wb.save("ImdbTop250.xlsx")
###Saved & Created xlsx file
print("File saved successfully")

